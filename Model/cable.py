# %%
import pandas as pd
import zipfile
import os
from datetime import datetime
from tqdm.auto import tqdm
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles import Font

# %%
def verify_zipped_netlist(netlist_file:str):
     if zipfile.is_zipfile(netlist_file):
          with zipfile.ZipFile(netlist_file, 'r') as zip_file:
               file_list = zip_file.namelist()
               # print(file_list)
               if 'pstxnet.dat' in file_list and 'pstxprt.dat' in file_list and 'pstxref.dat' in file_list:
                    # print('netlist is good')
                    return True
               else:
                    print("ERROR: Make sure pstxnet.dat, pstxprt.dat and pstxref.dat in zipped file.")
                    return False

     else:
          # print("\033[31mERROR: The configurations failed to operate correctly.\033[0m")
          return False

# print(verify_zipped_netlist("pstxprt.zip"))

# %%
class Board():
    def __init__(self, netlist_file):
        self.progress_bar = tqdm(total=4, bar_format="{l_bar}%s{bar}%s{r_bar}" % ('\033[36m', '\033[0m'))
        self.netlist_file = netlist_file
        self.progress_bar.update(1)
        self.board_name, self.export_physical_date = self.get_board_name_and_export_date()
        self.progress_bar.update(1)
        self.net_components_dict = self.get_nets_connection()
        self.progress_bar.update(1)
        self.__component_list = self.get_components_list()
        self.progress_bar.update(1)
        self.progress_bar.close()

    def get_board_name_and_export_date(self):
        ## Import pstxprt.dat to get board name & export physical time
        with zipfile.ZipFile(self.netlist_file, 'r') as archive:
            pstxprt_content = archive.read('pstxprt.dat')
            pstxprt_content = str(pstxprt_content)
            pstxprt_content = pstxprt_content.split('\\n')

            export_physical_time = pstxprt_content[5][12:-2]
            board_name = pstxprt_content[4][15:-2]
            # print(board_name)
            # print(export_physical_time)
            return board_name, export_physical_time

    def get_components_list(self):
        ## Not used in this program
        ## Import pstxprt.dat to get board name & export physical time
        with zipfile.ZipFile(self.netlist_file, 'r') as archive:
            pstxprt_content = archive.read('pstxprt.dat')
            pstxprt_content = str(pstxprt_content)
            pstxprt_content = pstxprt_content.split('\\n')

            all_components_list = []
            # print(pstxprt_content)
            for pstxprt_content_index, pstxprt_sub_content in enumerate(pstxprt_content):
                if 'PART_NAME' in pstxprt_sub_content and pstxprt_sub_content.index('PART_NAME') == 0:
                    all_components_list.append(pstxprt_content[pstxprt_content_index + 1][1:-1].split(' ')[0])
                    # print(pstxprt_content_index, pstxprt_sub_content) 
            # print(components_list)
            return all_components_list

    def get_nets_list(self):
        with zipfile.ZipFile(self.netlist_file, 'r') as archive:
            pstxnet_content = archive.read('pstxnet.dat')
            pstxnet_content = str(pstxnet_content)
            # print(pstxnet_netname)
            all_netname_list = []
            pstxnet_content = pstxnet_content.split('\\n')
            # print(pstxnet_content)

            for pstxnet_content_index, pstxnet_sub_content in enumerate(pstxnet_content):
                if 'NET_NAME' in pstxnet_sub_content and pstxnet_sub_content.index('NET_NAME') == 0:
                    # print(pstxnet_content_index, pstxnet_sub_content)
                    all_netname_list.append(pstxnet_content[pstxnet_content_index + 1][1:-1])
            # print(all_netname_list)

            return all_netname_list
    
    def is_component(self, __input_components_list:list):
        # __component_list = self.get_components_list()
        # print(__component_list)
        # print(__input_components_list)
        is_subset = set(__input_components_list) <= set(self.__component_list)
        if is_subset:
            # print("Legal component location")
            return True
        else:
            # print("ERROR: No such location(s).")
            return False
    
    
    def get_nets_connection(self):
        ## Import pstxref.dat 
        with zipfile.ZipFile(self.netlist_file, 'r') as archive:
            pstxref_content = archive.read('pstxref.dat')
            pstxref_content = str(pstxref_content)

            pstxref_content = pstxref_content.split('\\n\\n')
            # print(pstxref_content)

            ## Segment pstxref content only leaves the net-components section
            region_of_interest_start = 'END LOGICAL PART CROSS REFERENCE\\n\\x0cGLOBAL SIGNAL CROSS REFERENCE - ' + self.export_physical_date
            region_of_interest_index_start = 0
            region_of_interest_index_end = pstxref_content.index('END GLOBAL SIGNAL CROSS REFERENCE\\n\\x0cGLOBAL PART CROSS REFERENCE - ' + self.export_physical_date)
            # print(region_of_interest_index_end)
            # print(pstxref_content[418])

            for _region_of_interest_index_start, _pstxref_sub_content in enumerate(pstxref_content):
                if region_of_interest_start in _pstxref_sub_content and _pstxref_sub_content.index(region_of_interest_start) == 0:
                    # print(_region_of_interest_index_start)
                    region_of_interest_index_start = _region_of_interest_index_start
            region_of_interest_content = pstxref_content[region_of_interest_index_start:region_of_interest_index_end]
            # print(region_of_interest_content[0][len(region_of_interest_start) + 2:])
            region_of_interest_content[0] = region_of_interest_content[0][len(region_of_interest_start) + 2:]
            # print(region_of_interest_content)
            
            ## Build net_components connectivity dict
            single_net_list = []
            net_components_dict = {}
            for region_of_interest_content_index, region_of_interest_sub_content in enumerate(region_of_interest_content):
                    temp_content = region_of_interest_sub_content.split('\\n')  ## Make block a list for convinient
                    # print(temp_content[0].split(' ')[0])
                    net_components_dict[temp_content[0].split(' ')[0]] = []
                    for temp_content_index, temp_sub_content in enumerate(temp_content):
                        temp_sub_content = temp_sub_content.split(' ')
                        no_empty_temp_sub_content = [x for x in temp_sub_content if x]
                        if temp_content_index != 0:
                            # print(no_empty_temp_sub_content)
                            # print(no_empty_temp_sub_content[0] + '-' + no_empty_temp_sub_content[1])
                            net_components_dict[temp_content[0].split(' ')[0]].append(no_empty_temp_sub_content[0] + '-' + no_empty_temp_sub_content[1])
                            # print(net_components_dict[temp_content[0].split(' ')[0]])
                            _temp_list = ', '.join(map(str, net_components_dict[temp_content[0].split(' ')[0]]))
                    # print(_temp_list)
                    # net_components_dict[temp_content[0].split(' ')[0]] = _temp_list
            # # print(net_components_dict)
            return net_components_dict

    ## Consider if it necessary to trans component-list to compoonent_pin_dict
    ## It will be like component_pin_dict["J1"] --> {"J1-1", "J1-2", ..., "J1-N"}
    def get_component_pin_list(self, __component_location, excel_interconnection_format:bool = False):
        if __component_location in self.__component_list:
            with zipfile.ZipFile(self.netlist_file, 'r') as archive:
                pstxref_content = archive.read('pstxref.dat').decode('utf-8').split('\n\n')
            # print(len(pstxref_content))
            # print(pstxref_content)
            __component_location_format = __component_location.rjust(4) + ' '
            #print(__component_location_format)
            component_name_format_for_interconnection_excel_format = '  ' + __component_location + '\n'



            __component_pin_list = []

            if excel_interconnection_format:
                for pstxref_content_index, pstxref_content_element in enumerate(pstxref_content):
                    if component_name_format_for_interconnection_excel_format in pstxref_content_element:
                        content_sigment = pstxref_content_element.split('\n')[1:]  ## We don't want the first column
                        for content_sigment_sub_element in content_sigment:
                            content_sigment_sub_element = content_sigment_sub_element.split(' ')
                            content_sigment_sub_element_no_empty = [x for x in content_sigment_sub_element if x]
                            content_sigment_sub_element_no_empty[2], content_sigment_sub_element_no_empty[1] = content_sigment_sub_element_no_empty[1], content_sigment_sub_element_no_empty[2]
                            # print(content_sigment_sub_element_no_empty)
                            __component_pin_list.append(content_sigment_sub_element_no_empty[:3])
                # print(__component_pin_list)

            else:
                for pstxref_content_index, pstxref_content_element in enumerate(pstxref_content):
                    if pstxref_content_element.startswith(__component_location_format):
                        pstxref_content_element = pstxref_content_element.split('\n')
                        # print(pstxref_content_element)
        
                        for pstxref_content_sub_element in pstxref_content_element[1:]:
                            pstxref_content_sub_element_no_empty = [x for x in pstxref_content_sub_element.split(' ') if x]
                            # print(__component_location + '-' + pstxref_content_sub_element_no_empty[0])
                        
                            __component_pin_list.append(f"{__component_location}-{pstxref_content_sub_element_no_empty[0]}")
                            # if excel_interconnection_format:
                            #     __component_pin_list.append(pstxref_content_sub_element_no_empty[:2])
                            # else:
                                # __component_pin_list.append(f"{__component_location}-{pstxref_content_sub_element_no_empty[0]}")
            return __component_pin_list
        else:
            print(f"    ERROR: Entering component location not in this board!")
            return None

    def interconnection_path_mapping(self, __path_dict:dict):
        # print(mapping_dict)

        for component, component_path in __path_dict.items():
            # print(component)
            # Create mapping table by dict
            mapping_dict = {item[0]: item for item in self.get_component_pin_list(component, True)}  ## Into board class
            # print(mapping_dict)
            for component_pin_connectivity in component_path:
                # print(component_pin_connectivity)
                # print(component_pin_connectivity[0].split('-')[1])
                mapping_key = component_pin_connectivity[0].split('-')[1]
                if mapping_key in mapping_dict:
                    # print(component_pin_connectivity[:1])
                    # print(mapping_dict[mapping_key])
                    if len(component_pin_connectivity) == 1:
                        component_pin_connectivity[:1] = mapping_dict[mapping_key]
                        # print(component_pin_connectivity)
                    else:
                        # for i, j in enumerate(mapping_dict[mapping_key][:2]):
                        #     # print(i, j)
                        #     component_pin_connectivity.insert(i, j)
                        # component_pin_connectivity.pop(2)
                        temp = component_pin_connectivity[2:]
                        component_pin_connectivity.clear()
                        component_pin_connectivity.extend(mapping_dict[mapping_key]+temp)

# print(len(Board('pstprop.zip').get_component_pin_list('CPLD1', True)))
# Board('netlist/2u_fan_board.zip').get_component_pin_list('J10', True)
# Board('pstprop.zip').get_component_pin_list('U3', True)
# print(Board('E7142_2U_FAN_BOARD_NETLIST_20231029_0045.zip').get_component_pin_list('U3', True))

# %%
## Estimate the cable type (1-1, y cable or other shape)
class Cable():

    def __init__(self, cable_xlsx_file, DEBUG:bool = False):
        self.cable_file = cable_xlsx_file
        self.DEBUG = DEBUG
        self.df_cable = self.excel_number_to_string(self.cable_file)
        # print(self.df_cable)
        self.port_columns = self.df_cable.columns
        self.cable_file_name = os.path.basename(self.cable_file)
        self.number_of_port = self.get_number_of_port()
        self.df_cable_remove_pin_definition = self.get_functional_df_cable(self.df_cable)
        self.port_list = self.get_port_list()
        self.excel_pass_fail_condition = self.get_excel_pass_fail_condition()

    def debug_print(self, message):
        if self.DEBUG:
            print(f"DEBUG: {message}")

    def get_number_of_port(self):
        """
        Finds the maximum number in the list that matches the pattern ^P\d.
        
        Returns:
            int: The maximum number found, or None if no match is found.
        """
        # 用於匹配 ^P\d 的正則表達式
        pattern = r"^(P|p)(\d+)"
        max_number = None

        for item in self.df_cable.columns:
            match = re.match(pattern, item)
            self.debug_print(f'Match: {match}')
            if match:
                ## group(1) would be "P" or "p"
                ## group(2) would be (\d+)
                number = int(match.group(2))
                self.debug_print(f"Match number: {number}")
                # 更新最大值
                if max_number is None or number > max_number:
                    max_number = number

        self.debug_print(f"Number of port: {max_number}")
        return max_number
    
    def get_functional_df_cable(self, df_cable_await_to_modify):
        num_rows, num_columns =  self.df_cable.shape
        if self.number_of_port == num_columns:
            return df_cable_await_to_modify
        else:
            column_delete_list = []
            pattern = r"^(P|p)(\d+) (PIN|Pin|pin) (DEFINITION|Definition|definition)" 
            for index, item in enumerate(df_cable_await_to_modify.columns):
                match = re.match(pattern, item)
                if match:
                    column_delete_list.append(index)
            
            self.debug_print(f"index list: {column_delete_list}")
            df_cable_await_to_modify = df_cable_await_to_modify.drop(df_cable_await_to_modify.columns[column_delete_list], axis=1)
            # 動態生成列標題
            df_cable_await_to_modify.columns = [f'P{i+1}' for i in range(self.number_of_port)]
            self.debug_print(f"df_cable_await_to_modify: {df_cable_await_to_modify}")
            return df_cable_await_to_modify

    
    def get_port_list(self):
        """
        Extracts unique port identifiers (e.g., 'P1', 'P2', 'P3', ...) from the input port list.

        Args:
            port_list (list): A list of strings representing ports and their definitions.

        Returns:
            list: A list of unique port identifiers.
        """
        # Use a set to collect unique port identifiers
        ports = set()

        # Regular expression to match port identifiers like 'P1', 'P2', etc.
        pattern = r"^(P\d+)"
        
        for item in self.df_cable.columns:
            match = re.match(pattern, item)
            if match:
                ports.add(match.group(1))

        self.debug_print(f"Port list: {ports}")
        
        # Convert set to a sorted list (optional)
        return sorted(ports)
    
    def excel_number_to_string(self, __cable_file):
        cable_file = pd.read_excel(__cable_file)
        self.debug_print(f'Original excel data: {cable_file}')
        ## Trans numbers in df_cable to string
        for column in cable_file.columns:
            try: 
                cable_file[column] = cable_file[column].astype(pd.Int64Dtype())
                cable_file[column] = cable_file[column].astype(str)
            except:
                pass
        cable_file.replace(['<NA>', 'n', 'nan'], None, inplace=True)
        self.debug_print(f'Modified excel data: {cable_file}')
        return cable_file

    def generate_board_connection(self, **kwargs:dict):
        self.debug_print(f"Port clolumns: {self.df_cable.columns}")
        # Convert self.port_list to a list (if it's a Pandas Index)
        port_columns = list(self.df_cable.columns)

        # Iterate through the list and modify it
        for index, item in enumerate(port_columns):
            pattern = r"^(P|p)(\d+) (PIN|Pin|pin)(#| NUMBER| Number| number| No.| No)"
            match = re.match(pattern, item)
            if match:
                port_columns[index] = f"P{match.group(2)}"

        # Update self.port_columns with the modified list
        self.port_columns = port_columns
        self.debug_print(f'Modified port columns: {self.port_columns}')

        ## Get cable data from class cable itself
        # df_cable = self.df_cable_remove_pin_definition.copy()
        df_cable = self.df_cable.copy()
        df_cable.columns = port_columns
        self.debug_print(f"{kwargs.items()}")
        
        ## keys store cable port data; values store {netlist file}, {location}
        for key, value in kwargs.items():
            self.debug_print(f"key: {key}, value: {value}")
            self.debug_print(f"{key} Wired to board {value[0]}'s location {value[1]}")
            board = Board((value[0]))
            pin_list = board.get_component_pin_list(value[1], True)
            board_name = board.board_name
            df_pin_list = pd.DataFrame(pin_list, columns=[f'{board_name} {value[1]} Pin Number', 
                                                          f'{board_name} {value[1]} Pin Name', 
                                                          f'{board_name} {value[1]} Net Name'])
            
            # self.debug_print(f"DataFrame pin list: {df_pin_list}")

            for index, row in df_cable.iterrows():
                self.debug_print(f'index: {index}')
                self.debug_print(f'row: {row}')
                self.debug_print(f'!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
                # self.debug_print(f'row[key]: {row[key]}')
                if pd.notna(row[key]):
                    pin_number = row[key]            
                    # print(f'pin_number: {pin_number}')
                    # print(type(pin_number))  
                    # print(df_pin_list[f'{value[1]} Pin Number'])
                    if str(pin_number) in df_pin_list[f'{board_name} {value[1]} Pin Number'].values:
                        # print(df_pin_list[f'{value[1]} Pin Number'] == str(pin_number))
                        # print(df_pin_list[df_pin_list[f'{value[1]} Pin Number'] == str(pin_number)])
                        conn_row = df_pin_list[df_pin_list[f'{board_name} {value[1]} Pin Number'] == str(pin_number)].iloc[0]
                        # print(f'conn_row: {conn_row}')
                        df_cable.at[index, f'{board_name} {value[1]} Pin Number'] = conn_row[f'{board_name} {value[1]} Pin Number']
                        df_cable.at[index, f'{board_name} {value[1]} Pin Name'] = conn_row[f'{board_name} {value[1]} Pin Name']
                        df_cable.at[index, f'{board_name} {value[1]} Net Name'] = conn_row[f'{board_name} {value[1]} Net Name']
                    else:
                        df_cable.at[index, f'{board_name} {value[1]} Pin Number'] = f'nan'
                        df_cable.at[index, f'{board_name} {value[1]} Pin Name'] = f'nan'
                        df_cable.at[index, f'{board_name} {value[1]} Net Name'] = f'nan'
                        
                        print(f"\033[91mWarning: Cable {key} pin {pin_number} is not present in connector {value[1]}!\033[0m")
        ## Eliminate ['<NA>', 'n', 'nan'] on dataframe to make excel neat
        df_cable.replace(['<NA>', 'n', 'nan'], '', inplace=True)
        self.debug_print(f'RRRRRRRRRRRRRRRRRRRRRRRRRRR: {df_cable}')
        return(df_cable)
    
    def get_excel_pass_fail_condition(self):
        self.debug_print(f'Get excel data: {self.df_cable_remove_pin_definition}')
        bool_df_cable_data = self.df_cable_remove_pin_definition.notna()
        num_rows = bool_df_cable_data.shape[0]
        num_columns = bool_df_cable_data.shape[1]
        self.debug_print(f'Shape of bool_df: {bool_df_cable_data.shape}')
        self.debug_print(f'Excel bool map: {bool_df_cable_data}')

        ## Specific the column name instead of dynamic method
        new_column_names = []
        if num_columns == 2:
            new_column_names = ['G', 'J']
        elif num_columns ==3: 
            new_column_names = ['I', 'L', 'O']
        elif num_columns ==4:
            new_column_names = ['K', 'N', 'Q', 'T']

        bool_df_cable_data.columns = new_column_names

        formula_column = []
        row_offset = 3
        for row in bool_df_cable_data.itertuples(index=True):
            true_positions = [f"{col}{row.Index + row_offset}" for col in bool_df_cable_data.columns if getattr(row, col)]
            false_positions = [f"{col}{row.Index + row_offset}" for col in bool_df_cable_data.columns if not getattr(row, col)]

            # Construct the true condition output
            if true_positions:
                result_true = f'{true_positions[0]} <> ""'
                if len(true_positions) > 1:
                    result_true += ", " + ", ".join(f"{true_positions[0]} = {item}" for item in true_positions[1:])
            else:
                result_true = ''

            # Construct the false condition output
            result_false = ", ".join(f'{item} = ""' for item in false_positions)

            # Combine the results, removing any trailing commas or whitespace
            excel_condition = f"{result_true}, {result_false}".strip(", ")
            formula = f'=IF(AND({excel_condition}), "Pass", "Fail")'
            # Output the combined condition
            # print(formula)
            formula_column.append(formula)
        return formula_column

        



# %%
class Excel_former():

    def __init__(self, DEBUG:bool = False):
        self.DEBUG = DEBUG

    def debug_print(self, message):
        if self.DEBUG:
            print(f"DEBUG: {message}")

    def generate_cable_routing_report(self, df_cable_connection, board_name):
        # 格式化當前的日期與時間
        formatted_time = datetime.now().strftime("%Y%m%d_%H%M")
        self.df_cable_connection = df_cable_connection
        excel_output_path_and_name = board_name + '_' + f"{formatted_time}.xlsx"
        try:
            writer = pd.ExcelWriter(excel_output_path_and_name, engine='xlsxwriter')
            self.df_cable_connection.to_excel(writer, index = False, sheet_name=board_name)
            writer.close()
        except PermissionError: 
            print("\033[31mERROR: Please close [cable_connection_report.xlsx] file before running program\n")
        return excel_output_path_and_name
    
    def add_excel_pass_fail_condition(self, origin_cable_report, excel_condition_list):
        # Load raw excel report
        wb = load_workbook(origin_cable_report)
        ws = wb.active
        # print(excel_condition_list
        result_column = ws.max_column - 1
        target_column = get_column_letter(result_column)

        for row_num, value in enumerate(excel_condition_list, start = 3):
            ws[f'{target_column}{row_num}'] = value
        wb.save(origin_cable_report)



    def friendly_cable_report(self, origin_cable_report):
        # Remove duplicates while preserving order
        def remove_duplicates(input_list):
            seen = set()  # Create a set to track seen items
            output_list = []  # List to hold unique items
            
            for item in input_list:
                if item not in seen:  # If the item has not been seen
                    seen.add(item)  # Add to the set
                    output_list.append(item)  # Add to the output list
            
            return output_list


        # Load raw excel report
        wb = load_workbook(origin_cable_report)
        ws = wb.active
        num_columns = ws.max_column
        # print(num_columns)
        num_rows = ws.max_row
        # print(num_rows)
        # print(f'Number of total columns: {num_columns}')
        net_name_index = []

        ## The number of column should be {number_of_port}*5
        ## We can get number of cable port by num_columns / 5
        number_of_port = num_columns // 5
        self.debug_print(f'Number of port: {number_of_port}')
        
        # 提取第一行的所有儲存格值
        row_number = 1  # 替換為您想要讀取的行號（1 代表第一行）
        row_values = [cell.value for cell in ws[row_number]]
        self.debug_print(f'Row values: {row_values}')
        

        board_connector_list = [item.rsplit(' ', 2)[0] for item in row_values]
        self.debug_print(f'Board connector list: {board_connector_list}')
        pin_title_list = ['Pin Number', 'Pin Name', 'Net Name'] * number_of_port
        # print(pin_title_list)

        # Generate port list dynamically
        port_list = []
        for i in range(1, number_of_port + 1):
            port_list.extend([f'P{i} Pin Number', f'P{i} Pin Definition'])
        self.debug_print(f'Port List: {port_list}')
        second_row_title = port_list + pin_title_list
        self.debug_print(second_row_title)
        
        ## Insert a new row at the first position for showing category of cable and board
        ws.insert_rows(1)
        ## Dealing with the port column
        for port in range(1, number_of_port * 2, number_of_port *2 - 1):
            self.debug_print(f'Port in for loop: {port}')
            merge_range = f'{ws.cell(row=1, column=port).coordinate}:{ws.cell(row=1, column=number_of_port*2).coordinate}'
            self.debug_print(f'Merge range for cable port: {merge_range}')
            # Merge the specific cells
            ws.merge_cells(merge_range)
            # 設定合併後的儲存格文字和置中
            ws.cell(row=1, column=port).value = "Cable Port"  # 可替換成您想要的文字
            ws.cell(row=1, column=port).alignment = Alignment(horizontal='center', vertical='center')
            
        ## Dealing with the connector column
        for col in range(number_of_port * 2 + 1, number_of_port*2 + number_of_port * 3 + 1, 3):
            self.debug_print(f'col: {col}')
            # 設定要合併的範圍
            end_col = col + 2  # 合併三個欄位
            merge_range = f'{ws.cell(row=1, column=col).coordinate}:{ws.cell(row=1, column=end_col).coordinate}'
            self.debug_print(f'Merge range for board field: {merge_range}')
            
            # 合併儲存格
            ws.merge_cells(merge_range)
            
            # 設定合併後的儲存格文字和置中
            ws.cell(row=1, column=col).value = board_connector_list[col]  # 可替換成您想要的文字
            ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')


        ## Replace second row by neat title, 
        ## for instance: DCP_MB J1_HSBPPWR Pin Number, DCP_MB J1_HSBPPWR Pin Name, DCP_MB J1_HSBPPWR Pin Net Name, 
        ## to                              Pin Number,                   Pin Name,                   Pin Net Name
        for col_num, value in enumerate(second_row_title, start=1):
            self.debug_print(f'column number, value: {col_num}, {value}')
            ws.cell(row=2, column=col_num, value=value)
            if value == 'Net Name':
                net_name_index.append(col_num)
        self.debug_print(f'Net name index: {net_name_index}')
        

        # Add check result and commment for user checklist
        ws.cell(row=1, column=num_columns + 1, value='User Review Checklist')
        review_checklist_merge_range = f'{ws.cell(row=1, column=num_columns+1).coordinate}:{ws.cell(row=1, column=num_columns+2).coordinate}'
        # print(review_checklist_merge_range)
        ws.merge_cells(review_checklist_merge_range)
        ws.cell(row=2, column=num_columns + 1, value='Result')
        ws.cell(row=2, column=num_columns + 2, value='Comment')

        # ## For example: ['E', 'H'] for 2-port cable
        # def get_netname_column_letter(dynamic_netname_index_list:list):
        #     # Define the row you want to work with (example for row 3)
        #     col_vars = {}
            
        #     netname_index_list = []
        #     for index, value in enumerate(dynamic_netname_index_list):
        #         col_vars[f'netname_index_{index}'] = value  
        #         netname_index_list.append(get_column_letter(value))
        #     # print(f"col_vars: {col_vars}")
        #     # print(col_vars["netname_index_0"])
        #     # print(get_column_letter(col_vars["netname_index_0"]))

        #     return netname_index_list
        

        # netname_letter_list = get_netname_column_letter(net_name_index)
        # print(netname_letter_list)
        # print(len(netname_letter_list))

        # # formula = f"{netname_letter_list[0]}{num_rows}"
        # # print(formula)

        # ## In the case of a 2-port configuration
        # if len(netname_letter_list) == 2:
        #     for row in range(3, num_rows+2):
        #         formula = f'=IF(AND({netname_letter_list[0]}{row}<>"", {netname_letter_list[1]}{row}<>"", {netname_letter_list[0]}{row} = {netname_letter_list[1]}{row}), "Pass", "Fail")'
        #         # print(formula)
        #         ws[f'{get_column_letter(num_columns+1)}{row}'].value = formula

        # ## In the case of a 3-port configuration
        # elif len(netname_letter_list) == 3:
        #     for row in range(3, num_rows+2):
        #         formula = f'=IF(AND({netname_letter_list[0]}{row}<>"", OR(AND({netname_letter_list[0]}{row} = {netname_letter_list[1]}{row}, {netname_letter_list[2]}{row}=""), AND({netname_letter_list[0]}{row} = {netname_letter_list[2]}{row}, {netname_letter_list[1]}{row}=""))), "Pass", "Fail")'
        #         ws[f'{get_column_letter(num_columns+1)}{row}'].value = formula
        #         # print(formula)

        # ## In the case of a 4-port configuration
        # elif len(netname_letter_list) == 4:
        #     for row in range(3, num_rows+2):
        #         formula = f'=IF(AND({netname_letter_list[0]}{row}<>"", OR(AND({netname_letter_list[0]}{row} = {netname_letter_list[1]}{row}, {netname_letter_list[2]}{row}="", {netname_letter_list[3]}{row}=""), AND({netname_letter_list[0]}{row} = {netname_letter_list[2]}{row}, {netname_letter_list[1]}{row}="", {netname_letter_list[3]}{row}=""), AND({netname_letter_list[0]}{row} = {netname_letter_list[3]}{row}, {netname_letter_list[1]}{row}="", {netname_letter_list[2]}{row}=""))), "Pass", "Fail")'
        #         ws[f'{get_column_letter(num_columns+1)}{row}'].value = formula
        #         # print(formula)





        # 定義無邊框樣式
        no_border = Border(left=Side(style=None), 
                        right=Side(style=None), 
                        top=Side(style=None), 
                        bottom=Side(style=None))
        # Define the Segoe UI Historic font
        segoe_font = Font(name='Segoe UI Historic')

        # 遍歷整個工作表中的所有儲存格
        for row in ws.iter_rows():
            for cell in row:
                cell.border = no_border  # 設置為無邊框
                cell.font = segoe_font  # Set the font to Segoe UI Historic

        # 儲存修改後的 Excel 檔案
        wb.save(origin_cable_report)
        

# Excel_former().generate_cable_routing_report(cable.generate_board_connection(**test_dict), 'cable')

# %%
def main():  
    # print("\nCABLE CONNECTION TOOL LOG") 
    file_path = '../cable_data/test_cable.xlsx'
    # file_path = '../cable_data/1U_Fan_Board_cable.xlsx'
    # file_path = '../cable_data/2U_Fan_Board_cable.xlsx'
    # file_path = '../cable_data/new3.xlsx'
    # file_path = '../cable_data/2U_Fan_Board_cable_new.xlsx'

    cable = Cable(file_path, True)
    # cable = Cable(file_path)

    print(cable.excel_pass_fail_condition)
            

    test_dict = {'P1': ("../netlist/mb.zip", 'J1_HSBPPWR'),
                'P2': ('../netlist/E7142_1U_FAN_BOARD_R01.zip', 'J1'), 
                'P3': ('../netlist/2u_fan_board.zip', 'J1'),
                'P4': ('../netlist/E7142_CP_LEFT_INTERPOSER_R01_NETLIST_20240903_1052.zip', 'J2')}

    # test_dict = {'P1': ("../netlist/mb.zip", 'J1_HSBPPWR'),
    #             'P2': ('../netlist/2u_fan_board.zip', 'J1'), 
    #             'P3': ('../netlist/2u_fan_board.zip', 'J10')}

    # test_dict = {'P1': ("../netlist/mb.zip", 'J1_HSBPPWR'),
    #              'P2': ('../netlist/2u_fan_board.zip', 'J1')}
    
    # test_dict = {'P1': ("../netlist/mb.zip", 'J1_MXIO_SLOT3'),
    #             'P2': ('../netlist/2u_fan_board.zip', 'J10'), 
    #             'P3': ('../netlist/E7142_1U_FAN_BOARD_R01.zip', 'J1')}

    # print(cable.generate_board_connection(**test_dict))
    output_excel_file_name = Excel_former().generate_cable_routing_report(cable.generate_board_connection(**test_dict), 'cable')
    Excel_former().friendly_cable_report(output_excel_file_name)
    Excel_former().add_excel_pass_fail_condition(output_excel_file_name, cable.excel_pass_fail_condition)



if __name__ == "__main__":
    main()

# %%



